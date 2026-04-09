"""TRNT 급여 대시보드 — 사이드바 메뉴 방식 (Google Sheets DB)"""
import hashlib
import io
import json
import re
import secrets
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
import xlrd
import gspread
from google.oauth2.service_account import Credentials

# ──────────────────────────── Google Sheets 연결 ────────────────────────────
SPREADSHEET_ID = "1WiEk3DXYDavh54iXTinco4NJNzirrk3kRxSatx3nCIs"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

@st.cache_resource
def _get_gsheet():
    """Google Sheets 연결 (캐시)"""
    try:
        # Streamlit Cloud: secrets.toml에서 읽기
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    except Exception:
        # 로컬: JSON 파일에서 읽기
        cred_file = Path(__file__).parent / "trntpay-6be7aca8d7d7.json"
        creds = Credentials.from_service_account_file(str(cred_file), scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc.open_by_key(SPREADSHEET_ID)

# ──────────────────────────── 로컬 경로 (업로드 백업용) ────────────────────────────
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

COMPANIES = ["티알앤티", "트리니티 필라테스"]
RATE_FIELDS = ["개인", "듀엣", "자이로", "OT", "그룹", "본교육", "워크샵"]
TAX_FIELDS = ["국민연금", "건강보험", "고용보험", "장기요양", "소득세", "지방소득세", "공제합계"]

# ──────────────────────────── 유틸 ────────────────────────────
def _hash_pw(pw: str, salt: str) -> str:
    return hashlib.sha256((salt + pw).encode("utf-8")).hexdigest()

def won(n) -> str:
    return f"{int(round(float(n))):,}원"

def _num(v) -> float:
    try:
        return float(v) if v not in ("", None) else 0.0
    except (TypeError, ValueError):
        return 0.0

def _is_total(*vals) -> bool:
    for v in vals:
        s = str(v).replace(" ", "")
        if s.startswith("합계") or s.startswith("총계") or s.startswith("총합"):
            return True
    return False

def to_label(ym: str, multi_year: bool) -> str:
    y, m = ym.split("-")
    return f"{y[2:]}년 {int(m)}월" if multi_year else f"{int(m)}월"

# ──────────────────────────── Google Sheets 저장/로드 ────────────────────────────
def _gs_load(sheet_name, default=None):
    """Google Sheets 시트에서 A1 셀의 JSON 읽기"""
    try:
        sh = _get_gsheet()
        ws = sh.worksheet(sheet_name)
        val = ws.acell("A1").value
        if val:
            return json.loads(val)
    except Exception:
        pass
    return default if default is not None else {}

def _gs_save(sheet_name, data):
    """Google Sheets 시트의 A1 셀에 JSON 저장"""
    sh = _get_gsheet()
    ws = sh.worksheet(sheet_name)
    ws.update("A1", [[json.dumps(data, ensure_ascii=False)]])

def _gs_load_keyed(sheet_name, key, default=None):
    """시트 내에서 특정 key(월 등)로 저장된 데이터 로드. B열=key, C열=json"""
    try:
        sh = _get_gsheet()
        ws = sh.worksheet(sheet_name)
        records = ws.get_all_values()
        for row in records:
            if len(row) >= 2 and row[0] == key:
                return json.loads(row[1])
    except Exception:
        pass
    return default

def _gs_save_keyed(sheet_name, key, data):
    """시트 내에서 특정 key(월 등)로 데이터 저장 (upsert)"""
    sh = _get_gsheet()
    ws = sh.worksheet(sheet_name)
    records = ws.get_all_values()
    json_str = json.dumps(data, ensure_ascii=False)
    for i, row in enumerate(records):
        if len(row) >= 1 and row[0] == key:
            ws.update(f"A{i+1}", [[key, json_str]])
            return
    # 새 행 추가
    ws.append_row([key, json_str], value_input_option="RAW")

def load_settings():
    s = _gs_load("settings", {
        "subsidy_pretax": 1614254,
        "subsidy_posttax": 1445114,
        "instructor_rates": [],
        "admin_salary": [],
        "account_routing": [],
        "personnel": [],
        "other_payees": [],
        "cleaning_staff": {"name": "임자영", "hourly_rate": 12000},
    })
    s.setdefault("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
    return s

def save_settings(s): _gs_save("settings", s)

def load_smtp():
    cfg = _gs_load("smtp", {
        "host": "smtp.gmail.com", "port": 587, "user": "", "password": "",
        "sender_name": "TRNT", "manager_email": "",
    })
    cfg.setdefault("manager_email", cfg.pop("admin_email", ""))
    return cfg

def save_smtp(cfg): _gs_save("smtp", cfg)

def load_settlement(ym):
    return _gs_load_keyed("settlements", ym, None)

def save_settlement(ym, data):
    _gs_save_keyed("settlements", ym, data)

# ──────────────────────────── DB (급여대장) — Google Sheets ────────────────────────────
def load_db():
    records = _gs_load("payroll", [])
    return dedupe(records) if records else []

def save_db(records):
    _gs_save("payroll", dedupe(records))

def _load_auth():
    return _gs_load("auth", None)

def _save_auth(data):
    _gs_save("auth", data)

# ──────────────────────────── 급여대장 파싱 (기존) ────────────────────────────
def detect_meta(filename, sheet):
    name = unicodedata.normalize("NFC", filename)
    company = next((c for c in COMPANIES if c in name), "")
    kind = "사업소득" if "사업소득" in name else ("급여" if "급여대장" in name else "")
    ym = ""
    for r in range(min(sheet.nrows, 6)):
        for c in range(sheet.ncols):
            v = str(sheet.cell_value(r, c))
            m = re.search(r"(\d{4})\D+(\d{1,2})\s*월", v)
            if m:
                ym = f"{m.group(1)}-{int(m.group(2)):02d}"
                break
        if ym: break
    if not ym:
        m = re.search(r"(\d{1,2})\s*월", name)
        if m: ym = f"2026-{int(m.group(1)):02d}"
    return company, kind, ym

def parse_payroll(sheet):
    rows = []; r = 8
    while r < sheet.nrows:
        emp_no = str(sheet.cell_value(r, 0)).strip()
        name = str(sheet.cell_value(r, 1)).strip()
        if _is_total(emp_no, name): break
        if not emp_no: r += 1; continue
        if r + 2 >= sheet.nrows: break
        gross = _num(sheet.cell_value(r + 2, 8))
        pension = _num(sheet.cell_value(r, 9)); health = _num(sheet.cell_value(r, 10))
        employ = _num(sheet.cell_value(r, 11)); care = _num(sheet.cell_value(r, 12))
        inc_tax = _num(sheet.cell_value(r, 13)); local_tax = _num(sheet.cell_value(r, 14))
        ded = _num(sheet.cell_value(r + 1, 14))
        if ded == 0: ded = pension + health + employ + care + inc_tax + local_tax
        if name:
            rows.append({"name": name, "gross": gross, "type": "급여",
                "국민연금": pension, "건강보험": health, "고용보험": employ,
                "장기요양": care, "소득세": inc_tax, "지방소득세": local_tax, "공제합계": ded})
        r += 3
    return rows

def parse_business_income(sheet):
    rows = []; r = 6
    while r < sheet.nrows:
        no = str(sheet.cell_value(r, 0)).strip()
        name = str(sheet.cell_value(r, 2)).strip()
        if _is_total(no, name): break
        if not no: r += 1; continue
        gross = _num(sheet.cell_value(r, 4)); inc_tax = _num(sheet.cell_value(r, 5))
        local_tax = _num(sheet.cell_value(r + 1, 5)) if r + 1 < sheet.nrows else 0.0
        if name:
            rows.append({"name": name, "gross": gross, "type": "사업소득",
                "국민연금": 0, "건강보험": 0, "고용보험": 0, "장기요양": 0,
                "소득세": inc_tax, "지방소득세": local_tax, "공제합계": inc_tax + local_tax})
        r += 2
    return rows

def parse_xls_file(filename, file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sheet = wb.sheet_by_index(0)
    company, kind, ym = detect_meta(filename, sheet)
    if not (company and kind and ym):
        return None, f"메타데이터 추출 실패 (회사={company}, 종류={kind}, 년월={ym})"
    rows = parse_payroll(sheet) if kind == "급여" else parse_business_income(sheet)
    return {"company": company, "kind": kind, "ym": ym, "rows": rows}, None

# ──────────────────────────── 핏투데이 파일 파싱 ────────────────────────────
def parse_fittoday_xlsx(file_bytes, target_month=None):
    """핏투데이내역 .xlsx → {'fittoday': [...], 'academy': [...]}"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {"fittoday": [], "academy": []}

    # 핏투데이케어 시트
    if "핏투데이케어" in wb.sheetnames:
        ws = wb["핏투데이케어"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0] or "").strip().replace(" 선생님", "")
            if not name: continue
            result["fittoday"].append({
                "name": name,
                "개인_횟수": int(_num(row[1])), "개인_급여": int(_num(row[2])),
                "OT_횟수": int(_num(row[3])), "OT_급여": int(_num(row[4])),
                "듀엣_횟수": int(_num(row[5])), "듀엣_급여": int(_num(row[6])),
                "그룹_횟수": int(_num(row[7])), "그룹_급여": int(_num(row[8])),
            })

    # 스탓교육 월별 집계 시트
    # 컬럼: A=예시마크, B=년(2026년), C=월(3월), D=강사명, E=교육분류, F=교육명
    #        G=Lv1/Lv2/ISP 일수, H=시험/워크샵 시간, I=개인교육시간
    #        J=5인이상총시간, K=5인이상추가인원, L=듀엣시간, M=4인교육시간
    if "스탓교육 월별 집계" in wb.sheetnames:
        ws = wb["스탓교육 월별 집계"]
        for row in ws.iter_rows(min_row=9, values_only=True):  # row 9부터 (7=예시1, 8=예시2)
            year_raw = str(row[1] or "").strip()  # B열
            month_raw = str(row[2] or "").strip()  # C열
            name = str(row[3] or "").strip()       # D열
            if not name or not year_raw: continue
            # "2026년" → "2026"
            y_match = re.search(r"(\d{4})", year_raw)
            m_match = re.search(r"(\d+)", month_raw)
            if not y_match or not m_match: continue
            row_ym = f"{y_match.group(1)}-{int(m_match.group(1)):02d}"
            if target_month and row_ym != target_month: continue
            category = str(row[4] or "")   # E열
            detail = str(row[5] or "")     # F열
            result["academy"].append({
                "name": name, "category": category, "detail": detail,
                "lv1_days": int(_num(row[6])),      # G열: Lv1/Lv2/ISP 일수
                "exam_hours": int(_num(row[7])),     # H열: 시험/워크샵 시간
                "private_hours": int(_num(row[8])),  # I열: 개인교육 시간
                "group5_hours": int(_num(row[9])),   # J열: 5인이상 총시간
                "group5_extra": int(_num(row[10])),  # K열: 5인이상 추가인원
                "duet_hours": int(_num(row[11])),    # L열: 듀엣 시간
                "group4_hours": int(_num(row[12])),  # M열: 4인교육 시간
            })

    return result

# ──────────────────────────── 정산 계산 ────────────────────────────
def calculate_settlement(settings, fittoday_data, manual_inputs, ym):
    """1차(세전) + 2차(세금구분) 정산 계산"""
    rates = {r["name"]: r for r in settings.get("instructor_rates", [])}
    admins = {a["name"]: a for a in settings.get("admin_salary", [])}
    routing = {r["name"]: r for r in settings.get("account_routing", [])}
    result = []

    all_names = set()

    # 1) 핏투데이 레슨페이
    lesson_pay = {}
    for ft in fittoday_data.get("fittoday", []):
        name = ft["name"]
        all_names.add(name)
        rate = rates.get(name, {})
        pay = (
            ft["개인_횟수"] * rate.get("개인", 0)
            + ft["듀엣_횟수"] * rate.get("듀엣", 0)
            + ft["OT_횟수"] * rate.get("OT", 0)
            + ft["그룹_횟수"] * rate.get("그룹", 0)
        )
        # 핏투데이에서 이미 계산된 그룹/OT/듀엣 급여가 있으면 그걸 사용
        ft_pay = ft["개인_급여"] + ft["OT_급여"] + ft["듀엣_급여"] + ft["그룹_급여"]
        # 개인 급여만 단가로 계산 (핏투데이케어에서 개인급여는 항상 0)
        calculated_pay = (
            ft["개인_횟수"] * rate.get("개인", 0)
            + ft["OT_급여"]
            + ft["듀엣_급여"]
            + ft["그룹_급여"]
        )
        lesson_pay[name] = calculated_pay

    # 2) 아카데미
    academy_pay = {}
    for ac in fittoday_data.get("academy", []):
        name = ac["name"]
        all_names.add(name)
        rate = rates.get(name, {})
        edu_days = ac.get("lv1_days", 0)  # Lv1/Lv2/ISP 본교육 일수
        exam_hours = ac.get("exam_hours", 0)  # 시험/워크샵 시간
        # 본교육: 1일 = 5시간, 시급 × 5시간 / 워크샵·시험: 시급 × 시간
        pay = edu_days * rate.get("본교육", 0) * 5 + exam_hours * rate.get("워크샵", 0)
        academy_pay[name] = academy_pay.get(name, 0) + pay

    # 3) 관리자 급여
    admin_pay = {}
    for name, adm in admins.items():
        if adm.get("status") != "재직": continue
        all_names.add(name)
        # 퇴사자 일할 계산
        resign = manual_inputs.get("resignees", {}).get(name)
        if adm["pay_type"] == "시급":
            hours = manual_inputs.get("work_hours", {}).get(name, adm.get("work_hours", 0))
            base = adm.get("hourly_rate", 0) * hours
        else:
            base = adm.get("base_salary", 0)
        if resign:
            work_days = resign.get("work_days", 0)
            total_days = resign.get("total_days", 1)
            base = int(base * work_days / total_days) if total_days > 0 else 0
        admin_pay[name] = {
            "base": base,
            "non_taxable": adm.get("non_taxable", 0),
            "other_allowance": adm.get("other_allowance", 0),
        }

    # 4) 기타 급여 대상자
    other_pay = {}
    for op in manual_inputs.get("other_payees", settings.get("other_payees", [])):
        name = op["name"]
        all_names.add(name)
        other_pay[name] = int(op.get("amount", 0))

    # 5) 상여금/인센티브
    bonus_pay = {}
    for b in manual_inputs.get("bonuses", []):
        name = b["name"]
        all_names.add(name)
        bonus_pay[name] = bonus_pay.get(name, 0) + int(b.get("amount", 0))

    # 합산 → 1차 정산
    for name in sorted(all_names):
        adm = admin_pay.get(name, {})
        lp = lesson_pay.get(name, 0)
        ap = academy_pay.get(name, 0)
        op = other_pay.get(name, 0)
        bp = bonus_pay.get(name, 0)

        base_salary = adm.get("base", 0)
        non_tax = adm.get("non_taxable", 0)
        other_allow = adm.get("other_allowance", 0)
        lesson_total = lp + ap
        extra = op + bp

        gross = base_salary + non_tax + other_allow + lesson_total + extra

        # 2차 정산: 세금 구분
        admin_info = admins.get(name, {})
        route = routing.get(name, {})

        if base_salary > 0 and lesson_total > 0:
            tax_type = "기본급 초과분 3.3% 적용"
            salary_amount = base_salary + non_tax
            business_amount = lesson_total + extra + other_allow
        elif base_salary > 0:
            tax_type = "기본급 4대 보험 적용"
            salary_amount = base_salary + non_tax
            business_amount = extra + other_allow
        elif lesson_total > 0 or extra > 0:
            tax_type = "3.3% 원천세 적용"
            salary_amount = 0
            business_amount = lesson_total + extra + other_allow
        elif op > 0:
            tax_type = "3.3% 원천세 적용"
            salary_amount = 0
            business_amount = op
        else:
            tax_type = "대상없음"
            salary_amount = 0
            business_amount = 0

        result.append({
            "name": name,
            "base_salary": base_salary,
            "non_taxable": non_tax,
            "other_allowance": other_allow,
            "lesson_pay": lp,
            "academy_pay": ap,
            "lesson_total": lesson_total,
            "other_pay": op,
            "bonus": bp,
            "gross": gross,
            "tax_type": tax_type,
            "salary_amount": salary_amount,
            "business_amount": business_amount,
            "salary_company": route.get("salary_company", ""),
            "lesson_company": route.get("lesson_company", ""),
            "other_company": route.get("other_company", ""),
        })

    return result

def generate_accountant_excel(settlement_data, ym, settings):
    """세무사 전달용 Excel 1개 파일 (Sheet: 트리니티필라테스, 티알앤티)"""
    # 세무사 전달용 이름 변환 (동일 인물, 시스템명 다름)
    ACCOUNTANT_NAME_MAP = {"강다감": "강혜라"}

    personnel = {p["name"]: p for p in settings.get("personnel", [])}

    trinity_rows = []
    trnt_rows = []

    for item in settlement_data:
        name = item["name"]
        # 세무사 전달용 이름 변환
        display_name = ACCOUNTANT_NAME_MAP.get(name, name)
        # 주민번호는 변환된 이름으로도 조회
        pinfo = personnel.get(display_name, personnel.get(name, {}))
        rid = pinfo.get("resident_id", "")

        # 급여(4대보험) 배분
        if item["salary_amount"] > 0:
            company = item.get("salary_company", "")
            row = {"NO": "", "주민번호": rid, "이름": display_name, "구분": "급여",
                   "지급액": item["salary_amount"], "비고": ""}
            if "티알" in company or "아카데미" in company:
                trnt_rows.append(row)
            else:
                trinity_rows.append(row)

        # 사업소득 배분
        if item["business_amount"] > 0:
            lesson_co = item.get("lesson_company", "")
            row = {"NO": "", "주민번호": rid, "이름": display_name, "구분": "사업소득",
                   "지급액": item["business_amount"], "비고": ""}
            if "티알" in lesson_co or "아카데미" in lesson_co:
                trnt_rows.append(row)
            else:
                trinity_rows.append(row)

    # 번호 매기기
    for i, r in enumerate(trinity_rows, 1): r["NO"] = i
    for i, r in enumerate(trnt_rows, 1): r["NO"] = i

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        cols = ["NO", "주민번호", "이름", "구분", "지급액", "비고"]
        pd.DataFrame(trinity_rows, columns=cols).to_excel(
            writer, sheet_name="트리니티필라테스(센터)", index=False)
        pd.DataFrame(trnt_rows, columns=cols).to_excel(
            writer, sheet_name="티알앤티(아카데미)", index=False)
        # 지급액 컬럼에 천 단위 쉼표 서식 적용
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = '#,##0'
    return output.getvalue()

# ──────────────────────────── DB (기존 급여대장) ────────────────────────────
def dedupe(records):
    seen = {}
    for r in records:
        key = (r.get("company"), r.get("kind"), r.get("ym"), r.get("name"))
        seen[key] = r
    return list(seen.values())

def upsert(records, parsed):
    key = (parsed["company"], parsed["kind"], parsed["ym"])
    kept = [r for r in records if (r["company"], r["kind"], r["ym"]) != key]
    for row in parsed["rows"]:
        rec = {"company": parsed["company"], "kind": parsed["kind"],
               "ym": parsed["ym"], "name": row["name"], "gross": row["gross"]}
        for f in TAX_FIELDS:
            rec[f] = row.get(f, 0)
        kept.append(rec)
    return kept

def to_df(records):
    cols = ["company", "kind", "ym", "name", "gross"] + TAX_FIELDS
    if not records:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(records)
    for c in ["gross"] + TAX_FIELDS:
        if c not in df.columns: df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df

# ──────────────────────────── 이메일 ────────────────────────────
def build_payslip_html(name, ym, items):
    y, m = ym.split("-")
    title = f"[{int(m)}월 급여명세서] {name}님"
    total_gross = sum(it["gross"] for it in items)
    total_ded = sum(it["공제합계"] for it in items)
    total_net = total_gross - total_ded
    blocks = []
    for it in items:
        rows = [("지급 (세전)", it["gross"]), ("국민연금", it["국민연금"]),
                ("건강보험", it["건강보험"]), ("고용보험", it["고용보험"]),
                ("장기요양", it["장기요양"]), ("소득세", it["소득세"]),
                ("지방소득세", it["지방소득세"]), ("공제 합계", it["공제합계"]),
                ("실지급액", it["gross"] - it["공제합계"])]
        rows_html = "".join(
            f"<tr><td style='padding:6px 12px;border-bottom:1px solid #eee'>{k}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>{int(round(v)):,}원</td></tr>"
            for k, v in rows if v or k in ("지급 (세전)", "공제 합계", "실지급액"))
        blocks.append(
            f"<h3 style='margin:16px 0 8px;color:#333'>{it['company']} · {it['kind']}</h3>"
            f"<table style='border-collapse:collapse;width:100%;font-size:14px'>{rows_html}</table>")
    html = f"""<!doctype html><html><body style='font-family:-apple-system,Helvetica,sans-serif;max-width:560px;margin:0 auto;padding:24px;color:#222'>
<p style='font-size:16px;line-height:1.6'>{name}님, 안녕하세요.<br>이번달도 수고 하셨습니다.</p>
<p style='font-size:14px;color:#666'>{y}년 {int(m)}월 급여명세서를 보내드립니다.</p>
<hr style='border:none;border-top:1px solid #ddd;margin:16px 0'>
{''.join(blocks)}
<hr style='border:none;border-top:2px solid #333;margin:20px 0'>
<table style='border-collapse:collapse;width:100%;font-size:15px;font-weight:bold'>
<tr><td style='padding:6px 12px'>총 지급 (세전)</td><td style='padding:6px 12px;text-align:right'>{int(round(total_gross)):,}원</td></tr>
<tr><td style='padding:6px 12px;color:#c33'>총 공제</td><td style='padding:6px 12px;text-align:right;color:#c33'>-{int(round(total_ded)):,}원</td></tr>
<tr><td style='padding:6px 12px;color:#06c;font-size:17px'>실지급액</td><td style='padding:6px 12px;text-align:right;color:#06c;font-size:17px'>{int(round(total_net)):,}원</td></tr>
</table>
<p style='margin-top:24px;font-size:13px;color:#888'>— TRNT —</p></body></html>"""
    return title, html

def build_manager_summary_html(ym, mdf):
    y, m = ym.split("-")
    title = f"[{y}년 {int(m)}월 급여 집계] TRNT"
    items = ["국민연금", "건강보험", "고용보험", "장기요양", "소득세", "지방소득세", "공제합계"]
    def col_vals(sub):
        return [sub["gross"].sum()] + [sub[f].sum() for f in items] + [sub["gross"].sum() - sub["공제합계"].sum()]
    def hc(sub): return sub["name"].nunique()
    rows = ["지급액 (세전)"] + items + ["지급액 (세후)"]
    trnt_df = mdf[mdf["company"] == "티알앤티"]; trin_df = mdf[mdf["company"] == "트리니티 필라테스"]
    trnt = col_vals(trnt_df); trin = col_vals(trin_df); tot = col_vals(mdf)
    def fmt(v): return f"{int(round(v)):,}원"
    def td(v, bold=False, color=None):
        style = "padding:6px 12px;border-bottom:1px solid #eee;text-align:right"
        if bold: style += ";font-weight:bold"
        if color: style += f";color:{color}"
        return f"<td style='{style}'>{fmt(v)}</td>"
    header = ("<tr style='background:#f5f5f5'><th style='padding:8px 12px;text-align:left;border-bottom:2px solid #333'>항목</th>"
              "<th style='padding:8px 12px;text-align:right;border-bottom:2px solid #333'>티알앤티</th>"
              "<th style='padding:8px 12px;text-align:right;border-bottom:2px solid #333'>트리니티 필라테스</th>"
              "<th style='padding:8px 12px;text-align:right;border-bottom:2px solid #333'>합계</th></tr>")
    body_rows = []
    for i, label in enumerate(rows):
        bold = label in ("지급액 (세전)", "공제합계", "지급액 (세후)")
        color = "#06c" if label == "지급액 (세후)" else None
        body_rows.append(f"<tr><td style='padding:6px 12px;border-bottom:1px solid #eee{';font-weight:bold' if bold else ''}'>{label}</td>"
                         f"{td(trnt[i],bold,color)}{td(trin[i],bold,color)}{td(tot[i],bold,color)}</tr>")
    table_html = f"<table style='border-collapse:collapse;width:100%;font-size:14px'>{header}{''.join(body_rows)}</table>"
    html = f"""<!doctype html><html><body style='font-family:-apple-system,Helvetica,sans-serif;max-width:720px;margin:0 auto;padding:24px;color:#222'>
<h2 style='margin:0 0 4px'>📊 {y}년 {int(m)}월 급여 집계 요약</h2>
<p style='color:#666;margin:0 0 20px'>TRNT (티알앤티 · 트리니티 필라테스)</p>
<h3 style='margin:20px 0 8px;border-left:4px solid #06c;padding-left:10px'>회사별 세금·공제 집계</h3>
{table_html}
<table style='border-collapse:collapse;width:100%;font-size:14px;margin-top:16px'>
<tr style='background:#f5f5f5'><th style='padding:6px 12px;text-align:left;border-bottom:2px solid #333'>구분</th><th style='padding:6px 12px;text-align:right;border-bottom:2px solid #333'>인원</th></tr>
<tr><td style='padding:6px 12px;border-bottom:1px solid #eee'>티알앤티</td><td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>{hc(trnt_df)}명</td></tr>
<tr><td style='padding:6px 12px;border-bottom:1px solid #eee'>트리니티 필라테스</td><td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right'>{hc(trin_df)}명</td></tr>
<tr><td style='padding:6px 12px;border-bottom:1px solid #eee;font-weight:bold'>합계</td><td style='padding:6px 12px;border-bottom:1px solid #eee;text-align:right;font-weight:bold'>{hc(mdf)}명</td></tr>
</table>
<p style='margin-top:28px;font-size:12px;color:#888'>※ 본 요약에는 개인별 금액/명단이 포함되지 않습니다.</p></body></html>"""
    return title, html

def send_email(smtp_cfg, to_email, subject, html_body):
    import smtplib
    from email.mime.text import MIMEText
    from email.header import Header
    from email.utils import formataddr
    try:
        if isinstance(to_email, str):
            recipients = [e.strip() for e in to_email.split(",") if e.strip()]
        else:
            recipients = list(to_email)
        if not recipients: return False, "수신자 없음"
        msg = MIMEText(html_body, "html", "utf-8")
        msg["Subject"] = Header(subject, "utf-8")
        msg["From"] = formataddr((str(Header(smtp_cfg.get("sender_name") or "TRNT", "utf-8")), smtp_cfg["user"]))
        msg["To"] = ", ".join(recipients)
        with smtplib.SMTP(smtp_cfg["host"], int(smtp_cfg["port"])) as s:
            s.starttls()
            s.login(smtp_cfg["user"], smtp_cfg["password"])
            s.sendmail(smtp_cfg["user"], recipients, msg.as_string())
        return True, f"전송 완료 ({len(recipients)}명)"
    except Exception as e:
        return False, f"실패: {e}"

# ═══════════════════════════════════════════════════════════════════════
#                              페이지들
# ═══════════════════════════════════════════════════════════════════════

def save_monthly_input(ym, data):
    _gs_save_keyed("monthly_inputs", ym, data)

def load_monthly_input(ym):
    return _gs_load_keyed("monthly_inputs", ym, None)


def page_manager_input():
    """📤 매니저 입력 (로그인 불필요)"""
    st.header("📤 월간 급여 데이터 입력")
    st.caption("매니저/부원장이 매월 데이터를 입력하는 페이지입니다. 입력 후 '💾 저장' 버튼을 누르세요.")
    settings = load_settings()

    c1, c2 = st.columns(2)
    year = c1.selectbox("년도", [2025, 2026, 2027], index=1, key="mgr_year")
    month = c2.selectbox("월", list(range(1, 13)), index=2, key="mgr_month")
    ym = f"{year}-{month:02d}"

    # 기존 입력값 불러오기
    existing = load_monthly_input(ym)
    if existing:
        st.success(f"✅ {ym} 기존 입력값이 있습니다. 수정 후 다시 저장할 수 있습니다.")

    st.subheader("1️⃣ 핏투데이 파일 업로드")
    uploaded = st.file_uploader("핏투데이내역 .xlsx 파일", type=["xlsx"], key="mgr_upload")
    fittoday_data = {"fittoday": [], "academy": []}
    if uploaded:
        raw = uploaded.read()
        fittoday_data = parse_fittoday_xlsx(raw, target_month=ym)
        st.success(f"핏투데이케어: {len(fittoday_data['fittoday'])}명 / 스탓교육: {len(fittoday_data['academy'])}건")
        # 미리 세션에 저장
        st.session_state["mgr_fittoday"] = fittoday_data
        st.session_state["mgr_fittoday_raw"] = raw
    elif existing and existing.get("fittoday"):
        fittoday_data = existing["fittoday"]
        st.info(f"이전 업로드 데이터 사용 중 (핏투데이: {len(fittoday_data.get('fittoday',[]))}명)")
    elif "mgr_fittoday" in st.session_state:
        fittoday_data = st.session_state["mgr_fittoday"]

    st.subheader("2️⃣ 파트타임 관리자 근무시간")

    hourly_admins = [a for a in settings.get("admin_salary", []) if a.get("pay_type") == "시급" and a.get("status") == "재직"]
    ex_hours = existing.get("work_hours", {}) if existing else {}
    work_hours = {}
    if hourly_admins:
        cols = st.columns(len(hourly_admins))
        for i, adm in enumerate(hourly_admins):
            default_h = ex_hours.get(adm["name"], adm.get("work_hours", 0))
            work_hours[adm["name"]] = cols[i].number_input(
                f"{adm['name']} (시급 {adm.get('hourly_rate',0):,}원)",
                min_value=0, value=int(default_h), step=1, key=f"mgr_wh_{adm['name']}")
    else:
        st.caption("시급 직원이 없습니다.")

    st.subheader("3️⃣ 청소이모님 근무시간")
    cs = settings.get("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
    cs_name = cs.get("name", "임자영")
    cs_rate = int(cs.get("hourly_rate", 12000))
    st.caption(f"{cs_name}님 — 시급 {cs_rate:,}원")
    ex_cleaning = existing.get("cleaning_hours", 0) if existing else 0
    cleaning_hours = st.number_input("총 근무시간", min_value=0, value=int(ex_cleaning), step=1, key="mgr_cleaning")
    if cleaning_hours > 0:
        st.info(f"💰 예상 급여: {cleaning_hours * cs_rate:,}원 (시급 {cs_rate:,}원 × {cleaning_hours}시간)")

    st.divider()

    if st.button("💾 저장", type="primary"):
        input_data = {
            "ym": ym,
            "fittoday": fittoday_data,
            "work_hours": work_hours,
            "cleaning_hours": cleaning_hours,
        }
        # 원본 파일도 백업
        if "mgr_fittoday_raw" in st.session_state:
            (UPLOAD_DIR / f"{ym}_핏투데이내역.xlsx").write_bytes(st.session_state["mgr_fittoday_raw"])
        save_monthly_input(ym, input_data)
        st.success(f"✅ {ym} 데이터가 저장되었습니다. 관리자가 정산을 진행할 수 있습니다.")
        st.balloons()


def page_settlement():
    """📥 급여 정산 (관리자 전용)"""
    st.header("📥 급여 정산")
    settings = load_settings()

    c1, c2 = st.columns(2)
    year = c1.selectbox("년도", [2025, 2026, 2027], index=1, key="settle_year")
    month = c2.selectbox("월", list(range(1, 13)), index=2, key="settle_month")
    ym = f"{year}-{month:02d}"

    # 매니저 입력 데이터 불러오기
    mgr_input = load_monthly_input(ym)
    if not mgr_input:
        st.warning(f"⚠️ {ym} 매니저 입력 데이터가 없습니다. 매니저가 '📤 매니저 입력' 페이지에서 먼저 데이터를 입력해야 합니다.")
        # 저장된 입력 목록 표시
        try:
            sh = _get_gsheet()
            ws = sh.worksheet("monthly_inputs")
            all_vals = ws.get_all_values()
            saved_months = [row[0] for row in all_vals if row[0]]
            if saved_months:
                st.caption("저장된 월: " + ", ".join(saved_months))
        except Exception:
            pass
        return

    st.success(f"✅ {ym} 매니저 입력 데이터를 불러왔습니다.")

    # 입력값 요약 표시
    with st.expander("📋 매니저 입력 내역 확인", expanded=True):
        ft = mgr_input.get("fittoday", {})
        st.write(f"**핏투데이**: {len(ft.get('fittoday', []))}명 / 아카데미: {len(ft.get('academy', []))}건")
        wh = mgr_input.get("work_hours", {})
        if wh:
            st.write("**파트타임 근무시간**: " + ", ".join(f"{k}: {v}시간" for k, v in wh.items()))
        ops = mgr_input.get("other_payees", [])
        if ops:
            st.write("**기타 급여**: " + ", ".join(f"{o['name']}: {int(o['amount']):,}원" for o in ops))
        ch = mgr_input.get("cleaning_hours", 0)
        cs = settings.get("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
        if ch:
            st.write(f"**청소이모님 ({cs.get('name', '임자영')})**: {ch}시간 → {ch * int(cs.get('hourly_rate', 12000)):,}원")
        bns = mgr_input.get("bonuses", [])
        if bns:
            st.write("**상여금/인센티브**: " + ", ".join(f"{b['name']}: {int(b['amount']):,}원" for b in bns))

    # 관리자 전용 추가 입력
    st.subheader("추가 입력 (관리자 전용)")

    # 기타 급여 대상자 (기본 3명 + 추가/제외 가능, 최근 저장 유지)
    st.markdown("**기타 급여 대상자**")
    st.caption("기본 3명 외에 + 버튼으로 추가, 행 삭제로 제외 가능. 저장하면 다음 달에도 유지됩니다.")
    saved_others = settings.get("other_payees", [
        {"name": "정선아", "amount": 2500000, "note": ""},
        {"name": "임자영", "amount": 780000, "note": ""},
        {"name": "오경순", "amount": 1600000, "note": ""},
    ])
    # 청소이모님 근무시간이 매니저 입력에 있으면 미리 반영
    cs = settings.get("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
    cs_name = cs.get("name", "임자영")
    cs_rate = int(cs.get("hourly_rate", 12000))
    cleaning_hours = mgr_input.get("cleaning_hours", 0) if mgr_input else 0
    if cleaning_hours > 0:
        pay = cleaning_hours * cs_rate
        note = f"{cleaning_hours}시간 × {cs_rate:,}원 = {pay:,}원"
        found = False
        for o in saved_others:
            if o["name"] == cs_name:
                o["amount"] = pay
                o["note"] = note
                found = True
        if not found:
            saved_others.append({"name": cs_name, "amount": pay, "note": note})
    # None 비고 방지
    for o in saved_others:
        if o.get("note") is None:
            o["note"] = ""
    other_df = pd.DataFrame(saved_others if saved_others else [{"name": "", "amount": 0, "note": ""}])
    if "name" in other_df.columns:
        other_df = other_df.rename(columns={"name": "이름", "amount": "금액", "note": "비고"})
    edited_others = st.data_editor(other_df, num_rows="dynamic", key="settle_others_edit", use_container_width=True,
        column_config={"금액": st.column_config.NumberColumn(format="%d")})
    if st.button("💾 기타 급여 대상자 저장", key="save_others"):
        new_others = [{"name": r["이름"], "amount": int(r["금액"]), "note": str(r.get("비고", ""))}
                     for _, r in edited_others.iterrows() if r["이름"]]
        settings["other_payees"] = new_others
        save_settings(settings)
        st.success("저장 완료 — 다음 달에도 이 목록이 유지됩니다.")

    # 상여금/인센티브
    st.markdown("**상여금 / 인센티브 / 추가 급여**")
    bonus_df = pd.DataFrame([{"이름": "", "금액": 0, "사유": ""}])
    edited_bonus = st.data_editor(bonus_df, num_rows="dynamic", key="settle_bonus_edit", use_container_width=True,
        column_config={"금액": st.column_config.NumberColumn(format="%d")})

    # 퇴사자 일할 계산
    st.markdown("**퇴사자 일할 계산** (퇴사일만 입력하면 자동 계산)")
    import calendar
    total_days_in_month = calendar.monthrange(year, month)[1]
    active_admins = [a["name"] for a in settings.get("admin_salary", []) if a.get("status") == "재직" and a.get("base_salary", 0) > 0]
    resign_name = st.selectbox("퇴사자 이름", ["(해당 없음)"] + active_admins, key="resign_name")
    resign_info = {}
    if resign_name != "(해당 없음)":
        resign_date = st.date_input("퇴사일", value=date(year, month, 15), key="resign_date",
                                     min_value=date(year, month, 1), max_value=date(year, month, total_days_in_month))
        work_days = resign_date.day
        st.info(f"📅 {resign_name}: {month}월 1일 ~ {resign_date.day}일 근무 → **{work_days}일 / {total_days_in_month}일** ({work_days/total_days_in_month:.1%})")
        resign_info[resign_name] = {"work_days": work_days, "total_days": total_days_in_month}

    st.divider()

    if st.button("🔄 정산 실행", type="primary"):
        # 청소이모님 근무시간 → 기타 급여에 자동 반영
        cleaning_hours = mgr_input.get("cleaning_hours", 0)
        cs = settings.get("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
        cs_name = cs.get("name", "임자영")
        cs_rate = int(cs.get("hourly_rate", 12000))
        other_list = [{"name": r["이름"], "amount": int(r["금액"]), "note": str(r.get("비고", ""))}
                     for _, r in edited_others.iterrows() if r["이름"]]
        if cleaning_hours > 0:
            pay = cleaning_hours * cs_rate
            note = f"{cleaning_hours}시간 × {cs_rate:,}원 = {pay:,}원"
            found = False
            for o in other_list:
                if o["name"] == cs_name:
                    o["amount"] = pay
                    o["note"] = note
                    found = True
            if not found:
                other_list.append({"name": cs_name, "amount": pay, "note": note})

        manual_inputs = {
            "work_hours": mgr_input.get("work_hours", {}),
            "other_payees": other_list,
            "bonuses": [{"name": r["이름"], "amount": int(r["금액"]), "reason": str(r.get("사유", ""))}
                       for _, r in edited_bonus.iterrows() if r["이름"] and r["금액"]],
            "resignees": resign_info,
        }
        fittoday_data = mgr_input.get("fittoday", {"fittoday": [], "academy": []})
        settlement = calculate_settlement(settings, fittoday_data, manual_inputs, ym)
        settlement = [r for r in settlement if r["gross"] > 0]  # 0원 제외
        save_settlement(ym, {"ym": ym, "data": settlement, "manual": manual_inputs})
        st.session_state["current_settlement"] = settlement
        st.session_state["current_ym"] = ym
        st.success(f"✅ {ym} 정산 완료 — {len(settlement)}명")

    # 결과 표시
    settlement = st.session_state.get("current_settlement")
    if not settlement:
        # 이전에 저장된 정산 불러오기
        saved = load_settlement(ym)
        if saved:
            settlement = saved.get("data", [])
            st.session_state["current_settlement"] = settlement
            st.session_state["current_ym"] = ym

    if settlement:
        ym_disp = st.session_state.get("current_ym", ym)
        st.subheader(f"📊 {ym_disp} 1차 정산 (세전)")
        sdf = pd.DataFrame(settlement)
        display_cols = ["name", "base_salary", "lesson_pay", "academy_pay", "other_pay", "bonus", "gross", "tax_type"]
        display_names = {"name": "이름", "base_salary": "관리자급", "lesson_pay": "레슨페이",
                        "academy_pay": "아카데미", "other_pay": "기타", "bonus": "상여/인센",
                        "gross": "세전합계", "tax_type": "정산구분"}
        show = sdf[display_cols].rename(columns=display_names)
        for c in ["관리자급", "레슨페이", "아카데미", "기타", "상여/인센", "세전합계"]:
            if c in show.columns: show[c] = show[c].astype(int)
        st.dataframe(show.style.format({c: "{:,}" for c in ["관리자급", "레슨페이", "아카데미", "기타", "상여/인센", "세전합계"]}),
                     use_container_width=True, hide_index=True)

        st.subheader("📊 2차 정산 (세금 구분)")
        tax_cols = ["name", "tax_type", "salary_amount", "business_amount"]
        tax_names = {"name": "이름", "tax_type": "구분", "salary_amount": "급여(4대보험)", "business_amount": "사업소득(3.3%)"}
        tax_show = sdf[tax_cols].rename(columns=tax_names)
        for c in ["급여(4대보험)", "사업소득(3.3%)"]: tax_show[c] = tax_show[c].astype(int)
        st.dataframe(tax_show.style.format({c: "{:,}" for c in ["급여(4대보험)", "사업소득(3.3%)"]}),
                     use_container_width=True, hide_index=True)

        st.subheader("3️⃣ 세무사 전달용 다운로드")
        # 주민등록번호/이메일 누락 체크
        ACCOUNTANT_NAME_MAP = {"강다감": "강혜라"}
        personnel = {p["name"]: p for p in settings.get("personnel", [])}
        missing_rid = []
        missing_email = []
        for item in settlement:
            name = item["name"]
            display_name = ACCOUNTANT_NAME_MAP.get(name, name)
            pinfo = personnel.get(display_name, personnel.get(name, {}))
            rid = pinfo.get("resident_id", "")
            email = pinfo.get("email", "")
            if not rid:
                missing_rid.append(display_name)
            if not email:
                missing_email.append(display_name)

        if missing_rid:
            @st.dialog("⚠️ 주민등록번호 누락")
            def show_missing_popup():
                st.error(f"**주민등록번호 미등록 ({len(missing_rid)}명)**")
                for n in missing_rid:
                    st.write(f"- {n}")
                st.info("👉 **⚙️ 설정 → 👤 인적사항**에서 주민등록번호를 입력해 주세요.")

            if st.button("📥 세무사 전달용 Excel 다운로드", type="primary"):
                show_missing_popup()
        else:
            if missing_email:
                st.caption(f"ℹ️ 이메일 미등록 ({len(missing_email)}명): {', '.join(missing_email)}")
            excel_bytes = generate_accountant_excel(settlement, ym_disp, settings)
            ym_short = ym_disp.replace("-", "")[2:]
            st.download_button("📥 세무사 전달용 Excel 다운로드", excel_bytes,
                file_name=f"{ym_short}_TRNT트리니티필라테스_정산.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")


def page_monthly():
    """📅 월별 현황"""
    st.header("📅 월별 현황")
    records = load_db(); df = to_df(records)
    if df.empty:
        st.info("세무사에게 받은 급여대장/사업소득지급대장 .xls 파일을 '📦 업로드 이력' 메뉴에서 업로드해 주세요.")
        return
    months = sorted(df["ym"].unique())
    sel_ym = st.selectbox("조회할 월", months, index=len(months) - 1)
    mdf = df[df["ym"] == sel_ym]
    total = mdf["gross"].sum(); by_co = mdf.groupby("company")["gross"].sum()
    trnt = by_co.get("티알앤티", 0); trinity = by_co.get("트리니티 필라테스", 0)
    tax_total = mdf["공제합계"].sum(); net_total = total - tax_total

    c1, c2, c3 = st.columns(3)
    c1.metric("총 급여액 (세전)", won(total)); c2.metric("티알앤티", won(trnt)); c3.metric("트리니티 필라테스", won(trinity))
    c4, c5, c6 = st.columns(3)
    c4.metric("세금·공제 합계", won(tax_total)); c5.metric("실지급액 (세후)", won(net_total))
    c6.metric("원천세 (소득세+지방)", won(mdf["소득세"].sum() + mdf["지방소득세"].sum()))

    with st.expander("세금·공제 상세 내역", expanded=False):
        items = ["국민연금", "건강보험", "고용보험", "장기요양", "소득세", "지방소득세", "공제합계"]
        rows_list = ["지급액 (세전)"] + items + ["지급액 (세후)"]
        def col_for(sub_df):
            return [sub_df["gross"].sum()] + [sub_df[f].sum() for f in items] + [sub_df["gross"].sum() - sub_df["공제합계"].sum()]
        tax_bd = pd.DataFrame({"항목": rows_list,
            "티알앤티": col_for(mdf[mdf["company"] == "티알앤티"]),
            "트리니티 필라테스": col_for(mdf[mdf["company"] == "트리니티 필라테스"]),
            "합계": col_for(mdf)})
        st.dataframe(tax_bd.style.format({"티알앤티": "{:,.0f}", "트리니티 필라테스": "{:,.0f}", "합계": "{:,.0f}"}),
                     use_container_width=True, hide_index=True)

    st.subheader("사람별 급여")
    person = mdf.groupby(["company", "name", "kind"])[["gross", "공제합계"]].sum().reset_index()
    person["실지급액"] = person["gross"] - person["공제합계"]
    person = person.sort_values(["company", "gross"], ascending=[True, False])
    person = person.rename(columns={"company": "회사", "name": "이름", "kind": "구분", "gross": "지급액(세전)", "공제합계": "공제"})
    for c in ["지급액(세전)", "공제", "실지급액"]: person[c] = person[c].astype(int)
    st.dataframe(person.style.format({"지급액(세전)": "{:,}", "공제": "{:,}", "실지급액": "{:,}"}),
                 use_container_width=True, hide_index=True)

    chart_metric = st.radio("그래프 기준", ["지급액(세전)", "실지급액"], horizontal=True)
    fig = px.bar(person, x="이름", y=chart_metric, color="회사", barmode="group", title=f"{sel_ym} 사람별 {chart_metric}")
    fig.update_yaxes(tickformat=",.0f", ticksuffix="원")
    fig.update_traces(hovertemplate="이름=%{x}<br>" + chart_metric + "=%{y:,.0f}원<extra></extra>")
    st.plotly_chart(fig, use_container_width=True)


def page_trends():
    """📈 추이 분석"""
    st.header("📈 추이 분석")
    records = load_db(); df = to_df(records)
    if df.empty: st.info("데이터가 없습니다."); return
    months = sorted(df["ym"].unique())
    multi_year = len({m.split("-")[0] for m in months}) > 1
    month_order = [to_label(m, multi_year) for m in months]
    won_axis = dict(tickformat=",.0f", ticksuffix="원")

    mt = df.groupby("ym")["gross"].sum().reset_index()
    mt["월"] = mt["ym"].apply(lambda x: to_label(x, multi_year)); mt = mt.rename(columns={"gross": "총급여"})
    fig1 = px.line(mt, x="월", y="총급여", markers=True, title="총 급여 추이", category_orders={"월": month_order})
    fig1.update_xaxes(type="category"); fig1.update_yaxes(**won_axis)
    fig1.update_traces(hovertemplate="월=%{x}<br>총급여=%{y:,.0f}원<extra></extra>")
    st.plotly_chart(fig1, use_container_width=True)

    mc = df.groupby(["ym", "company"])["gross"].sum().reset_index()
    mc["월"] = mc["ym"].apply(lambda x: to_label(x, multi_year)); mc = mc.rename(columns={"company": "회사", "gross": "급여"})
    fig2 = px.line(mc, x="월", y="급여", color="회사", markers=True, title="사업장별 급여 추이", category_orders={"월": month_order})
    fig2.update_xaxes(type="category"); fig2.update_yaxes(**won_axis)
    fig2.update_traces(hovertemplate="월=%{x}<br>급여=%{y:,.0f}원<extra></extra>")
    st.plotly_chart(fig2, use_container_width=True)

    st.subheader("사람별 급여 추이")
    co_filter = st.multiselect("회사 필터", COMPANIES, default=COMPANIES)
    pdf = df[df["company"].isin(co_filter)]
    names = sorted(pdf["name"].unique())
    name_filter = st.multiselect("이름 필터 (비우면 전체)", names)
    if name_filter: pdf = pdf[pdf["name"].isin(name_filter)]
    mp = pdf.groupby(["ym", "name"])["gross"].sum().reset_index()
    mp["월"] = mp["ym"].apply(lambda x: to_label(x, multi_year)); mp = mp.rename(columns={"name": "이름", "gross": "급여"})
    fig3 = px.line(mp, x="월", y="급여", color="이름", markers=True, title="사람별 급여 추이", category_orders={"월": month_order})
    fig3.update_xaxes(type="category"); fig3.update_yaxes(**won_axis)
    fig3.update_traces(hovertemplate="이름=%{fullData.name}<br>월=%{x}<br>급여=%{y:,.0f}원<extra></extra>")
    st.plotly_chart(fig3, use_container_width=True)


def page_payslip():
    """📧 급여명세서"""
    st.header("📧 급여명세서")
    records = load_db(); df = to_df(records)
    if df.empty: st.info("급여대장 데이터가 없습니다."); return
    months = sorted(df["ym"].unique())
    settings = load_settings(); smtp_cfg = load_smtp()

    with st.expander("✉️ SMTP 설정", expanded=False):
        c1, c2 = st.columns(2)
        smtp_cfg["host"] = c1.text_input("SMTP 호스트", smtp_cfg.get("host", "smtp.gmail.com"))
        smtp_cfg["port"] = c2.number_input("포트", value=int(smtp_cfg.get("port", 587)), step=1)
        smtp_cfg["user"] = st.text_input("발송 계정 (이메일)", smtp_cfg.get("user", ""))
        smtp_cfg["password"] = st.text_input("비밀번호 (앱 비밀번호 권장)", smtp_cfg.get("password", ""), type="password")
        smtp_cfg["sender_name"] = st.text_input("발신자 이름", smtp_cfg.get("sender_name", "TRNT"))
        smtp_cfg["manager_email"] = st.text_input("매니저 이메일", smtp_cfg.get("manager_email", ""))
        if st.button("SMTP 설정 저장"): save_smtp(smtp_cfg); st.success("저장됨")

    sel_ym = st.selectbox("발송할 월", months, index=len(months) - 1, key="slip_ym")
    mdf = df[df["ym"] == sel_ym]
    # 이름 매핑 (강다감 → 강혜라 이메일 사용)
    PAYSLIP_EMAIL_MAP = {"강다감": "강혜라"}
    personnel = {p["name"]: p for p in settings.get("personnel", [])}
    persons = sorted(mdf["name"].unique())

    def _get_email(name):
        mapped = PAYSLIP_EMAIL_MAP.get(name, name)
        return personnel.get(mapped, personnel.get(name, {})).get("email", "")

    missing = [p for p in persons if not _get_email(p)]
    if missing: st.warning(f"이메일 미등록 ({len(missing)}명): {', '.join(missing)} → '⚙️ 설정' 메뉴에서 등록하세요.")

    for person in persons:
        items = mdf[mdf["name"] == person].to_dict("records")
        email = _get_email(person)
        title, html = build_payslip_html(person, sel_ym, items)
        gross = sum(it["gross"] for it in items); ded = sum(it["공제합계"] for it in items)
        with st.expander(f"{person} · 세전 {won(gross)} · 실지급 {won(gross - ded)} · {email or '❌ 미등록'}"):
            st.components.v1.html(html, height=520, scrolling=True)
            if email and smtp_cfg.get("user") and smtp_cfg.get("password"):
                if st.button(f"📤 {person}에게 발송", key=f"send_{person}"):
                    ok, msg = send_email(smtp_cfg, email, title, html)
                    (st.success if ok else st.error)(msg)

    st.divider()
    # 매니저 요약
    manager_to = smtp_cfg.get("manager_email", "")
    st.markdown("### 매니저에게 월간 집계 발송")
    st.caption("**개인별 금액·이름은 포함되지 않습니다.**")
    if manager_to:
        with st.expander("발송될 집계 미리보기"):
            _, preview = build_manager_summary_html(sel_ym, mdf)
            st.components.v1.html(preview, height=400, scrolling=True)
        if st.button("📤 매니저에게 발송", key="send_mgr"):
            title, html = build_manager_summary_html(sel_ym, mdf)
            ok, msg = send_email(smtp_cfg, manager_to, title, html)
            (st.success if ok else st.error)(msg)

    st.divider()
    can_send = bool(smtp_cfg.get("user") and smtp_cfg.get("password"))
    if st.button("📤 전체 발송 (강사 개별 명세서)", type="primary", disabled=not can_send):
        prog = st.progress(0.0); log = []
        sendable = [p for p in persons if _get_email(p)]
        for i, p in enumerate(sendable, 1):
            items = mdf[mdf["name"] == p].to_dict("records")
            title, html = build_payslip_html(p, sel_ym, items)
            ok, msg = send_email(smtp_cfg, _get_email(p), title, html)
            log.append(f"{'✅' if ok else '❌'} {p}: {msg}"); prog.progress(i / len(sendable))
        st.write("\n\n".join(log))


def page_upload():
    """📦 업로드 이력 (급여대장/사업소득지급대장)"""
    st.header("📦 급여대장 업로드 & 이력")
    st.caption("세무사에게 받은 .xls 파일(급여대장/사업소득지급대장)을 업로드하세요.")

    uploads = st.file_uploader("엑셀(.xls) 파일 선택", type=["xls"], accept_multiple_files=True)
    if uploads:
        records = load_db(); msgs = []
        for up in uploads:
            raw = up.read(); parsed, err = parse_xls_file(up.name, raw)
            if err: msgs.append(("error", f"{up.name}: {err}")); continue
            records = upsert(records, parsed)
            archive = f"{parsed['ym']}_{parsed['company']}_{parsed['kind']}.xls"
            (UPLOAD_DIR / archive).write_bytes(raw)
            msgs.append(("ok", f"{up.name} → {parsed['company']} / {parsed['kind']} / {parsed['ym']} ({len(parsed['rows'])}명)"))
        save_db(records); df = to_df(records)
        for level, m in msgs:
            (st.success if level == "ok" else st.error)(m)

    records = load_db(); df = to_df(records)
    if not df.empty:
        st.subheader("저장된 데이터")
        summary = df.groupby(["ym", "company", "kind"]).agg(인원=("name", "nunique"), 세전합계=("gross", "sum")).reset_index()
        summary["세전합계"] = summary["세전합계"].astype(int)
        summary = summary.rename(columns={"ym": "월", "company": "회사", "kind": "종류"}).sort_values(["월", "회사"])
        st.dataframe(summary.style.format({"세전합계": "{:,}"}), use_container_width=True, hide_index=True)

    st.subheader("백업된 원본 파일")
    files = sorted(UPLOAD_DIR.glob("*.xls"))
    if not files: st.info("백업 파일 없음")
    else:
        for f in files:
            cols = st.columns([4, 1, 1])
            cols[0].write(f"📄 **{f.name}**")
            cols[1].caption(f"{f.stat().st_size / 1024:,.0f} KB")
            cols[2].download_button("⬇️", f.read_bytes(), file_name=f.name, key=f"dl_{f.name}")

    st.subheader("특정 월 삭제")
    months = sorted(df["ym"].unique()) if not df.empty else []
    if months:
        del_ym = st.selectbox("삭제할 월", months, key="del_ym")
        if st.button(f"🗑️ {del_ym} 삭제"):
            records = [r for r in records if r["ym"] != del_ym]
            save_db(records)
            for f in UPLOAD_DIR.glob(f"{del_ym}_*.xls"): f.unlink()
            st.success("삭제 완료"); st.rerun()


def page_raw():
    """🗂 원본 데이터"""
    st.header("🗂 원본 데이터")
    records = load_db(); df = to_df(records)
    if df.empty: st.info("데이터 없음"); return
    st.dataframe(df.sort_values(["ym", "company", "kind", "name"]), use_container_width=True, hide_index=True)
    st.download_button("CSV 다운로드", df.to_csv(index=False).encode("utf-8-sig"), file_name="payroll_all.csv", mime="text/csv")


def page_settings():
    """⚙️ 설정"""
    st.header("⚙️ 설정")
    settings = load_settings()

    tab1, tab2, tab3, tab_clean, tab4, tab5 = st.tabs(["👤 인적사항", "💰 강사 단가표", "🏢 관리자 급여", "🧹 청소이모님", "🔀 계좌 배정", "🔑 비밀번호"])

    with tab1:
        st.subheader("👤 인적사항 (이름 / 주민등록번호 / 이메일)")
        st.caption("신규 입사자는 하단 + 버튼으로 추가, 퇴사자는 행 삭제 대신 '⚙️ 관리자 급여/단가표'에서 재직상태를 '퇴사'로 변경하세요.")
        pdata = settings.get("personnel", [])
        pdf = pd.DataFrame(pdata if pdata else [{"name": "", "resident_id": "", "email": ""}])
        pdf = pdf[["name", "resident_id", "email"]]
        pdf.columns = ["이름", "주민등록번호", "이메일"]
        edited = st.data_editor(pdf, num_rows="dynamic", key="personnel_edit", use_container_width=True)
        if st.button("💾 인적사항 저장", type="primary"):
            settings["personnel"] = [{"name": r["이름"], "resident_id": r["주민등록번호"], "email": r["이메일"]}
                                     for _, r in edited.iterrows() if r["이름"]]
            save_settings(settings); st.success("저장 완료"); st.rerun()

    with tab2:
        st.subheader("💰 강사 단가표")
        st.caption("신규 강사: + 버튼으로 추가 / 퇴사: 재직상태를 '퇴사'로 변경")
        irates = settings.get("instructor_rates", [])
        idf = pd.DataFrame(irates if irates else [{"name":"","role":"","status":"재직","insurance":False,"개인":0,"듀엣":0,"자이로":0,"OT":0,"그룹":0,"본교육":0,"워크샵":0}])
        idf = idf[["name", "role", "status", "insurance"] + RATE_FIELDS]
        idf.columns = ["이름", "구분", "재직상태", "4대보험"] + RATE_FIELDS
        edited_i = st.data_editor(idf, num_rows="dynamic", key="rates_edit", use_container_width=True,
            column_config={"재직상태": st.column_config.SelectboxColumn(options=["재직", "퇴사"]),
                          "4대보험": st.column_config.CheckboxColumn()})
        if st.button("💾 단가표 저장", type="primary"):
            new_rates = []
            for _, r in edited_i.iterrows():
                if not r["이름"]: continue
                new_rates.append({"name": r["이름"], "role": r["구분"], "status": r["재직상태"],
                    "insurance": bool(r["4대보험"]),
                    **{f: int(r[f]) for f in RATE_FIELDS}})
            settings["instructor_rates"] = new_rates; save_settings(settings); st.success("저장 완료"); st.rerun()

    with tab3:
        st.subheader("🏢 관리자 급여 설정")
        st.caption("시급 직원은 유형='시급' / 기본급 직원은 유형='기본급'으로 설정")
        admins = settings.get("admin_salary", [])
        adf = pd.DataFrame(admins if admins else [{"name":"","type":"","status":"재직","pay_type":"기본급","base_salary":0,"hourly_rate":0,"insurance_rate":0,"non_taxable":0,"subsidy":False,"subsidy_end":""}])
        cols_show = ["name", "type", "status", "pay_type", "base_salary", "hourly_rate", "insurance_rate", "non_taxable", "subsidy", "subsidy_end"]
        adf = adf[[c for c in cols_show if c in adf.columns]]
        adf.columns = ["이름", "구분", "상태", "유형", "기본급", "시급", "보험적용률", "비과세", "지원금", "지원금종료"]
        edited_a = st.data_editor(adf, num_rows="dynamic", key="admin_edit", use_container_width=True,
            column_config={"상태": st.column_config.SelectboxColumn(options=["재직", "퇴사"]),
                          "유형": st.column_config.SelectboxColumn(options=["기본급", "시급"]),
                          "지원금": st.column_config.CheckboxColumn()})
        c1, c2 = st.columns(2)
        settings["subsidy_pretax"] = c1.number_input("지원금 세전", value=settings.get("subsidy_pretax", 1614254), step=1)
        settings["subsidy_posttax"] = c2.number_input("지원금 세후", value=settings.get("subsidy_posttax", 1445114), step=1)
        if st.button("💾 관리자 급여 저장", type="primary"):
            new_admins = []
            for _, r in edited_a.iterrows():
                if not r["이름"]: continue
                new_admins.append({"name": r["이름"], "type": r["구분"], "status": r["상태"], "pay_type": r["유형"],
                    "base_salary": int(r["기본급"]), "hourly_rate": int(r["시급"]),
                    "insurance_rate": float(r["보험적용률"]), "non_taxable": int(r["비과세"]),
                    "subsidy": bool(r["지원금"]), "subsidy_end": str(r["지원금종료"])})
            settings["admin_salary"] = new_admins; save_settings(settings); st.success("저장 완료"); st.rerun()

    with tab_clean:
        st.subheader("🧹 청소이모님 설정")
        st.caption("사람이 바뀌면 이름과 시급을 수정하세요. 매니저 입력 페이지에 자동 반영됩니다.")
        cs = settings.get("cleaning_staff", {"name": "임자영", "hourly_rate": 12000})
        cs_name = st.text_input("이름", value=cs.get("name", "임자영"), key="cs_name")
        cs_rate = st.number_input("시급 (원)", value=int(cs.get("hourly_rate", 12000)), step=1000, key="cs_rate")
        if st.button("💾 청소이모님 설정 저장", type="primary"):
            settings["cleaning_staff"] = {"name": cs_name, "hourly_rate": cs_rate}
            save_settings(settings)
            st.success(f"저장 완료 — {cs_name} (시급 {cs_rate:,}원)")
            st.rerun()

    with tab4:
        st.subheader("🔀 급여 지급 계좌 배정")
        st.caption("이름별로 급여·사업소득이 어느 회사(트리니티/티알앤티)로 지급되는지 설정")
        routing = settings.get("account_routing", [])
        rdf = pd.DataFrame(routing if routing else [{"name":"","salary_company":"","lesson_company":"","other_company":"","note":""}])
        rdf = rdf[["name", "salary_company", "lesson_company", "other_company", "note"]]
        rdf.columns = ["이름", "급여(4대보험)", "사업소득(레슨)", "사업소득(기타)", "비고"]
        edited_r = st.data_editor(rdf, num_rows="dynamic", key="routing_edit", use_container_width=True)
        if st.button("💾 계좌 배정 저장", type="primary"):
            new_routing = [{"name": r["이름"], "salary_company": r["급여(4대보험)"],
                "lesson_company": r["사업소득(레슨)"], "other_company": r["사업소득(기타)"], "note": r["비고"]}
                for _, r in edited_r.iterrows() if r["이름"]]
            settings["account_routing"] = new_routing; save_settings(settings); st.success("저장 완료"); st.rerun()

    with tab5:
        st.subheader("🔑 관리자 비밀번호 변경")
        old = st.text_input("현재 비밀번호", type="password", key="pw_old")
        new1 = st.text_input("새 비밀번호", type="password", key="pw_ch1")
        new2 = st.text_input("새 비밀번호 확인", type="password", key="pw_ch2")
        if st.button("변경"):
            auth = _load_auth()
            if not auth or _hash_pw(old, auth["salt"]) != auth["hash"]:
                st.error("현재 비밀번호가 맞지 않습니다.")
            elif not new1 or len(new1) < 4: st.error("최소 4자 이상")
            elif new1 != new2: st.error("일치하지 않습니다.")
            else:
                salt = secrets.token_hex(16)
                _save_auth({"salt": salt, "hash": _hash_pw(new1, salt)})
                st.success("변경 완료")


# ═══════════════════════════════════════════════════════════════════════
#                              메인
# ═══════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="TRNT 급여 대시보드", layout="wide")

# 사이드바 메뉴 (매니저 입력은 항상 보임)
with st.sidebar:
    st.title("TRNT 급여")

    ALL_PAGES = {
        "📤 매니저 입력": {"fn": page_manager_input, "login": False},
        "📥 급여 정산": {"fn": page_settlement, "login": True},
        "📅 월별 현황": {"fn": page_monthly, "login": True},
        "📈 추이 분석": {"fn": page_trends, "login": True},
        "📧 급여명세서": {"fn": page_payslip, "login": True},
        "📦 급여대장 업로드": {"fn": page_upload, "login": True},
        "🗂 원본 데이터": {"fn": page_raw, "login": True},
        "⚙️ 설정": {"fn": page_settings, "login": True},
    }

    page = st.radio("메뉴", list(ALL_PAGES.keys()), label_visibility="collapsed")

    st.divider()
    is_authed = st.session_state.get("authed", False)
    if is_authed:
        if st.button("🔒 로그아웃", use_container_width=True):
            st.session_state["authed"] = False; st.rerun()
    else:
        st.caption("🔐 관리자 메뉴는 로그인이 필요합니다")

# 페이지 렌더링
selected = ALL_PAGES[page]
if selected["login"] and not st.session_state.get("authed"):
    # 로그인 필요한 페이지 → 로그인 화면 표시
    auth = _load_auth()
    st.title("🔐 관리자 로그인")
    st.info(f"**'{page}'** 메뉴는 관리자 로그인이 필요합니다.")
    if not auth:
        st.warning("최초 접속입니다. 관리자 비밀번호를 설정해 주세요.")
        pw1 = st.text_input("새 비밀번호", type="password", key="pw_new1")
        pw2 = st.text_input("비밀번호 확인", type="password", key="pw_new2")
        if st.button("비밀번호 설정", type="primary"):
            if not pw1 or len(pw1) < 4: st.error("최소 4자 이상")
            elif pw1 != pw2: st.error("일치하지 않습니다.")
            else:
                salt = secrets.token_hex(16)
                _save_auth({"salt": salt, "hash": _hash_pw(pw1, salt)})
                st.session_state["authed"] = True; st.rerun()
    else:
        pw = st.text_input("관리자 비밀번호", type="password", key="pw_login")
        if st.button("로그인", type="primary"):
            if _hash_pw(pw, auth["salt"]) == auth["hash"]:
                st.session_state["authed"] = True; st.rerun()
            else: st.error("비밀번호가 맞지 않습니다.")
else:
    selected["fn"]()
